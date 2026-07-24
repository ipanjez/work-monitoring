import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\Farhans-WINDOWS\OneDrive - m365\Learn\Coding\Dashboard Proyek Soda Ash\Template_Dashboard_v3.xlsx')

def set_col(ws, header_row, data_rows, col_idx, header_text, default_text):
    ws.cell(row=header_row, column=col_idx).value = header_text
    ws.cell(row=header_row, column=col_idx).font = openpyxl.styles.Font(bold=True)
    for r in data_rows:
        ws.cell(row=r, column=col_idx).value = default_text

# 1. Config Sheet
wsC = wb['Config']
justification_text = '[Isi referensi sumber bobot, misal: Keputusan Manajemen / Standar Industri]'

# PHS Weights (Col 4 / D)
set_col(wsC, 11, range(12, 17), 4, 'Justifikasi Bobot', justification_text)
wsC.column_dimensions['D'].width = 40

# Overall Risk Index Weights (Col 4 / D)
set_col(wsC, 20, range(21, 24), 4, 'Justifikasi Bobot', justification_text)

# 2. Referensi_Threshold Sheet
wsR = wb['Referensi_Threshold']

# SRI (Col 6 / F, since C:E are merged)
set_col(wsR, 18, range(19, 25), 6, 'Justifikasi Bobot', justification_text)
# PEI
set_col(wsR, 27, range(28, 33), 6, 'Justifikasi Bobot', justification_text)
# CRI
set_col(wsR, 35, range(36, 41), 6, 'Justifikasi Bobot', justification_text)

wsR.column_dimensions['F'].width = 40

wb.save(r'c:\Users\Farhans-WINDOWS\OneDrive - m365\Learn\Coding\Dashboard Proyek Soda Ash\Template_Dashboard_v3.xlsx')
print('Excel justification columns added successfully')
